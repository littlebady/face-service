from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import List
from urllib.parse import quote
import re

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import PatternFill
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore[assignment]
    BarChart = None  # type: ignore[assignment]
    Reference = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]


router = APIRouter(tags=["数据分析"])

_INTEGER_COLUMN_NAMES = {"total_checkin", "revoked_checkin", "choice_question"}


def _decode_txt(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="ignore")


def _split_line_keep_all(line: str) -> List[str]:
    return [item.strip() for item in line.split("\t\t")]


def _split_line_remove_first_two(line: str) -> List[str]:
    cells = _split_line_keep_all(line)
    if len(cells) < 2:
        raise ValueError("TXT 列数不足，无法按旧格式解析。")
    return cells[2:]


def _safe_int(text: str, *, default: int = 0) -> int:
    try:
        return int(text)
    except (TypeError, ValueError):
        return default


def _normalize_row(raw: List[str], expected_len: int) -> List[str]:
    if len(raw) == expected_len:
        return raw
    if len(raw) > expected_len:
        return raw[:expected_len]
    return raw + [""] * (expected_len - len(raw))


def _build_workbook_from_txt(content: bytes):
    if Workbook is None or BarChart is None or Reference is None or PatternFill is None:
        raise HTTPException(status_code=500, detail="openpyxl 未安装，请先执行 pip install openpyxl")

    text = _decode_txt(content)
    lines = [line.rstrip("\r") for line in text.split("\n") if line.strip()]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "签到情况表"
    orange_fill = PatternFill(fill_type="solid", fgColor="FFA500")

    if not lines:
        return workbook

    new_format = "total_checkin" in lines[0]

    if new_format:
        header = _split_line_keep_all(lines[0])
        if len(header) < 2:
            raise HTTPException(status_code=400, detail="表头格式异常：缺少学号/姓名列。")
        header[0] = "学号"
        header[1] = "姓名"

        integer_cols = [False] * len(header)
        short_answer_seen = 0
        for idx, name in enumerate(header):
            if name in _INTEGER_COLUMN_NAMES:
                integer_cols[idx] = True
            elif name == "short_answer_question":
                short_answer_seen += 1
                if short_answer_seen == 1:
                    integer_cols[idx] = True

        for col_idx, value in enumerate(header, start=1):
            sheet.cell(row=1, column=col_idx, value=value)

        for row_idx, line in enumerate(lines[1:], start=2):
            row_values = _normalize_row(_split_line_keep_all(line), len(header))
            for col_idx, raw in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if integer_cols[col_idx - 1]:
                    cell.value = _safe_int(raw)
                else:
                    cell.value = raw
                if raw == "0":
                    cell.fill = orange_fill

        total_col_idx = next((i for i, h in enumerate(header) if h == "total_checkin"), -1)
        if total_col_idx < 0:
            raise HTTPException(status_code=400, detail="新格式中未找到 total_checkin 列。")
    else:
        header = _split_line_remove_first_two(lines[0])
        if len(header) < 3:
            raise HTTPException(status_code=400, detail="旧格式表头列数不足。")
        header[0] = "学号"
        header[1] = "姓名"
        header[-1] = "总计"

        for col_idx, value in enumerate(header, start=1):
            sheet.cell(row=1, column=col_idx, value=value)

        for row_idx, line in enumerate(lines[1:], start=2):
            row_values = _normalize_row(_split_line_remove_first_two(line), len(header))
            for col_idx, raw in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if col_idx == len(header):
                    cell.value = _safe_int(raw)
                else:
                    cell.value = raw
                if raw == "0":
                    cell.fill = orange_fill

        total_col_idx = len(header) - 1

    row_count = max(0, len(lines) - 1)
    if row_count > 0:
        chart = BarChart()
        chart.title = "签到人员和签到次数"
        chart.y_axis.title = "次数"
        chart.x_axis.title = "姓名"
        chart.width = 14
        chart.height = 8

        values = Reference(
            sheet,
            min_col=total_col_idx + 1,
            min_row=1,
            max_col=total_col_idx + 1,
            max_row=row_count + 1,
        )
        categories = Reference(
            sheet,
            min_col=2,
            min_row=2,
            max_col=2,
            max_row=row_count + 1,
        )
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        chart.legend.position = "t"
        sheet.add_chart(chart, f"A{row_count + 4}")

    return workbook


def _build_download_name(original_name: str | None) -> str:
    if not original_name:
        return "output.xlsx"
    path = Path(original_name)
    stem = path.stem if path.suffix else path.name
    return f"{stem}.xlsx"


def _build_ascii_fallback_name(filename: str) -> str:
    stem = Path(filename).stem
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-")
    if not safe_stem:
        safe_stem = "output"
    return f"{safe_stem}.xlsx"


@router.post(
    "/api/excel/generate",
    summary="生成签到 Excel",
    description="上传签到 TXT 数据并返回 Excel 文件（兼容旧格式与新格式）。",
)
async def generate_excel(file: UploadFile = File(..., description="签到明细 TXT 文件")):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件为空。")

    workbook = _build_workbook_from_txt(content)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    download_name = _build_download_name(file.filename)
    ascii_name = _build_ascii_fallback_name(download_name)
    encoded_name = quote(download_name, safe="")
    headers = {
        "Content-Disposition": f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
